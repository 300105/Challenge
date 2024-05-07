import openpyxl

def procurar_palavra_em_excel(excel_file, palavra):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(excel_file)
    
    # Iterar sobre todas as planilhas do arquivo
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Iterar sobre todas as células da planilha
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar se a célula contém a palavra
                if palavra.lower() in str(cell.value).lower():
                    print("A palavra '{}' foi encontrada na planilha '{}'.".format(palavra, sheet_name))
                    return

    print("A palavra '{}' não foi encontrada no arquivo Excel.".format(palavra))

# Exemplo de uso
excel_file = "execel 1.xlsx"
palavra = "TELEFONE"
procurar_palavra_em_excel(excel_file,palavra)